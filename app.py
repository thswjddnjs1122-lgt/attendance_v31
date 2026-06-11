from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, flash, send_file
import sqlite3, os, base64, uuid
from datetime import datetime
from io import BytesIO
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'static', 'uploads')
SIGN_DIR = os.path.join(BASE_DIR, 'static', 'signatures')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(SIGN_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = 'attendance-v32-secret-key-change-later'

ADMIN_ID = 'admin'
ADMIN_PW = '1234'


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_column(cur, table, column, col_type='TEXT'):
    cur.execute(f"PRAGMA table_info({table})")
    columns = [row[1] for row in cur.fetchall()]
    if column not in columns:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            department TEXT,
            grade TEXT,
            student_name TEXT,
            student_id TEXT,
            birth TEXT,
            created_at TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            department TEXT,
            grade TEXT,
            student_name TEXT,
            student_id TEXT,
            birth TEXT,
            semester TEXT,
            evidence_type TEXT,
            diagnosis_text TEXT,
            absence_period TEXT,
            reason TEXT,
            apply_date TEXT,
            student_signature TEXT,
            parent_name TEXT,
            parent_signature TEXT,
            attachment TEXT,
            ocr_text TEXT,
            auto_check TEXT,
            status TEXT DEFAULT '대기',
            reject_reason TEXT,
            created_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    ''')
    for col in ['parent_name', 'semester', 'evidence_type', 'diagnosis_text', 'ocr_text', 'auto_check']:
        ensure_column(cur, 'applications', col, 'TEXT')
    conn.commit()
    conn.close()


def save_signature(data_url, prefix):
    if not data_url or ',' not in data_url:
        return ''
    header, encoded = data_url.split(',', 1)
    try:
        image_data = base64.b64decode(encoded)
        filename = f"{prefix}_{uuid.uuid4().hex}.png"
        path = os.path.join(SIGN_DIR, filename)
        with open(path, 'wb') as f:
            f.write(image_data)
        return filename
    except Exception:
        return ''


def current_user():
    if 'user_id' not in session:
        return None
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    conn.close()
    return user


def guess_semester(apply_date_text):
    try:
        year = ''.join(ch for ch in apply_date_text[:4] if ch.isdigit()) or str(datetime.now().year)
        month_digits = ''.join(ch for ch in apply_date_text[5:8] if ch.isdigit())
        month = int(month_digits) if month_digits else datetime.now().month
        term = '1학기' if 3 <= month <= 8 else '2학기'
        return f'{year}-{term}'
    except Exception:
        now = datetime.now()
        term = '1학기' if 3 <= now.month <= 8 else '2학기'
        return f'{now.year}-{term}'


def run_ocr_if_possible(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ['.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff']:
        return 'OCR 생략: 이미지 파일이 아니거나 PDF/HWP 파일입니다.'
    try:
        from PIL import Image
        import pytesseract
        text = pytesseract.image_to_string(Image.open(file_path), lang='kor+eng')
        return (text or '').strip() or 'OCR 결과 없음'
    except Exception as e:
        return f'OCR 실패: {e}'


def make_auto_check(reason, diagnosis_text, ocr_text):
    reason = (reason or '').strip()
    diagnosis_text = (diagnosis_text or '').strip()
    ocr_text = (ocr_text or '').strip()
    notes = []
    if diagnosis_text:
        if diagnosis_text in reason or reason in diagnosis_text:
            notes.append('결석사유/진단내용 일치 가능')
        elif diagnosis_text in ocr_text or reason in ocr_text:
            notes.append('증빙서류 OCR에서 관련 문구 확인 가능')
        else:
            notes.append('주의: 결석사유와 진단내용이 다를 수 있음')
    else:
        notes.append('주의: 진단내용/병명 미입력')

    if ocr_text.startswith('OCR 실패'):
        notes.append('OCR 확인 필요')
    elif ocr_text.startswith('OCR 생략'):
        notes.append('OCR 생략: 관리자 수동 확인')
    elif ocr_text and ocr_text != 'OCR 결과 없음':
        notes.append('OCR 실행 완료')
    return ' / '.join(notes)


@app.route('/')
def index():
    if session.get('admin'):
        return redirect(url_for('admin_dashboard'))
    if session.get('user_id'):
        return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    user = current_user()
    if not user:
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_pw = request.form.get('current_pw', '')
        new_pw = request.form.get('new_pw', '')
        new_pw_confirm = request.form.get('new_pw_confirm', '')

        if not current_pw or not new_pw:
            flash('현재 비밀번호와 새 비밀번호를 모두 입력해주세요.')
            return redirect(url_for('change_password'))

        if new_pw_confirm and new_pw != new_pw_confirm:
            flash('새 비밀번호 확인이 일치하지 않습니다.')
            return redirect(url_for('change_password'))

        if not check_password_hash(user['password'], current_pw):
            flash('현재 비밀번호가 일치하지 않습니다.')
            return redirect(url_for('change_password'))

        conn = get_db()
        conn.execute(
            'UPDATE users SET password=? WHERE id=?',
            (generate_password_hash(new_pw), user['id'])
        )
        conn.commit()
        conn.close()

        flash('비밀번호가 변경되었습니다. 다시 로그인해주세요.')
        session.clear()
        return redirect(url_for('login'))

    return render_template('change_password.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        department = request.form.get('department', '').strip()
        grade = request.form.get('grade', '').strip()
        student_name = request.form.get('student_name', '').strip()
        student_id = request.form.get('student_id', '').strip()
        birth = request.form.get('birth', '').strip()
        if not username or not password or not student_name:
            flash('아이디, 비밀번호, 성명은 필수입니다.')
            return redirect(url_for('register'))
        conn = get_db()
        try:
            conn.execute('''INSERT INTO users(username,password,department,grade,student_name,student_id,birth,created_at)
                         VALUES(?,?,?,?,?,?,?,?)''',
                         (username, generate_password_hash(password), department, grade, student_name, student_id, birth, datetime.now().strftime('%Y-%m-%d %H:%M')))
            conn.commit()
        except sqlite3.IntegrityError:
            flash('이미 사용 중인 아이디입니다.')
            return redirect(url_for('register'))
        finally:
            conn.close()
        flash('회원가입 완료. 로그인해주세요.')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username == ADMIN_ID and password == ADMIN_PW:
            session.clear()
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password'], password):
            session.clear()
            session['user_id'] = user['id']
            return redirect(url_for('student_dashboard'))
        flash('아이디 또는 비밀번호가 올바르지 않습니다.')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/dashboard')
def student_dashboard():
    user = current_user()
    if not user:
        return redirect(url_for('login'))
    conn = get_db()
    apps = conn.execute('SELECT * FROM applications WHERE user_id=? ORDER BY id DESC', (user['id'],)).fetchall()
    conn.close()
    return render_template('dashboard.html', user=user, apps=apps)


@app.route('/apply', methods=['GET', 'POST'])
def apply():
    user = current_user()
    if not user:
        return redirect(url_for('login'))

    today = datetime.now().strftime('%Y년 %m월 %d일')
    default_semester = guess_semester(today)

    if request.method == 'POST':
        semester = request.form.get('semester') or default_semester
        conn = get_db()
        count = conn.execute('SELECT COUNT(*) AS cnt FROM applications WHERE user_id=? AND semester=?', (user['id'], semester)).fetchone()['cnt']
        if count >= 3:
            conn.close()
            flash(f'{semester} 신청 가능 횟수는 최대 3회입니다.')
            return redirect(url_for('student_dashboard'))

        attachment_name = ''
        ocr_text = ''
        file = request.files.get('attachment')
        if file and file.filename:
            safe_name = secure_filename(file.filename)
            attachment_name = f"{uuid.uuid4().hex}_{safe_name}"
            file_path = os.path.join(UPLOAD_DIR, attachment_name)
            file.save(file_path)
            ocr_text = run_ocr_if_possible(file_path)

        student_sig = save_signature(request.form.get('student_signature'), 'student')
        parent_sig = save_signature(request.form.get('parent_signature'), 'parent')
        reason = request.form.get('reason')
        diagnosis_text = request.form.get('diagnosis_text')
        auto_check = make_auto_check(reason, diagnosis_text, ocr_text)

        conn.execute('''INSERT INTO applications(
            user_id, department, grade, student_name, student_id, birth,
            semester, evidence_type, diagnosis_text,
            absence_period, reason, apply_date, student_signature, parent_name, parent_signature,
            attachment, ocr_text, auto_check, status, created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            user['id'], request.form.get('department'), request.form.get('grade'),
            request.form.get('student_name'), request.form.get('student_id'), request.form.get('birth'),
            semester, request.form.get('evidence_type'), diagnosis_text,
            request.form.get('absence_period'), reason, request.form.get('apply_date'),
            student_sig, request.form.get('parent_name'), parent_sig,
            attachment_name, ocr_text, auto_check, '대기', datetime.now().strftime('%Y-%m-%d %H:%M')
        ))
        conn.commit()
        conn.close()
        flash('신청이 저장되었습니다.')
        return redirect(url_for('student_dashboard'))

    return render_template('apply.html', user=user, today=today, default_semester=default_semester)


@app.route('/print/<int:app_id>')
def print_form(app_id):
    if not session.get('admin') and not session.get('user_id'):
        return redirect(url_for('login'))
    conn = get_db()
    app_row = conn.execute('SELECT * FROM applications WHERE id=?', (app_id,)).fetchone()
    conn.close()
    if not app_row:
        flash('신청서를 찾을 수 없습니다.')
        return redirect(url_for('index'))
    if session.get('user_id') and app_row['user_id'] != session.get('user_id'):
        flash('본인 신청서만 열람할 수 있습니다.')
        return redirect(url_for('student_dashboard'))
    return render_template('print_form.html', app=app_row)


@app.route('/admin')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('login'))
    status_filter = request.args.get('status', '전체')
    conn = get_db()
    if status_filter in ['대기', '승인', '반려']:
        apps = conn.execute('SELECT * FROM applications WHERE status=? ORDER BY student_name COLLATE NOCASE ASC, id DESC', (status_filter,)).fetchall()
    else:
        apps = conn.execute('SELECT * FROM applications ORDER BY student_name COLLATE NOCASE ASC, id DESC').fetchall()
    conn.close()
    return render_template('admin.html', apps=apps, status_filter=status_filter)


@app.route('/admin/status/<int:app_id>', methods=['POST'])
def update_status(app_id):
    if not session.get('admin'):
        return redirect(url_for('login'))
    status = request.form.get('status')
    reject_reason = request.form.get('reject_reason', '')
    conn = get_db()
    conn.execute('UPDATE applications SET status=?, reject_reason=? WHERE id=?', (status, reject_reason, app_id))
    conn.commit()
    conn.close()
    flash('상태가 변경되었습니다.')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/excel')
def admin_excel():
    if not session.get('admin'):
        return redirect(url_for('login'))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return 'openpyxl이 설치되어 있지 않습니다. CMD에서 pip install openpyxl 실행 후 다시 시도하세요.'

    conn = get_db()
    rows = conn.execute('SELECT * FROM applications ORDER BY student_name COLLATE NOCASE ASC, id DESC').fetchall()
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    headers = ['번호', '이름', '학번', '학과', '학년', '학기', '결석기간', '결석사유', '증빙종류', '진단내용', '자동검사', '상태', '반려사유', '신청일']

    def add_sheet(title, data):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for r in data:
            ws.append([r['id'], r['student_name'], r['student_id'], r['department'], r['grade'], r['semester'], r['absence_period'], r['reason'], r['evidence_type'], r['diagnosis_text'], r['auto_check'], r['status'], r['reject_reason'], r['created_at']])
        header_fill = PatternFill('solid', fgColor='D9EAF7')
        thin = Side(style='thin', color='999999')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        widths = [8, 12, 15, 18, 8, 14, 24, 28, 16, 20, 36, 10, 24, 18]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = 'A2'

    all_rows = list(rows)
    add_sheet('전체명단', all_rows)
    add_sheet('승인명단', [r for r in all_rows if r['status'] == '승인'])
    add_sheet('반려명단', [r for r in all_rows if r['status'] == '반려'])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"attendance_list_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    if not session.get('admin') and not session.get('user_id'):
        return redirect(url_for('login'))
    return send_from_directory(UPLOAD_DIR, filename)


if __name__ == '__main__':
    init_db()
    app.run(debug=True)
